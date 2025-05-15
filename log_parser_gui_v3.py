# ✅ Full Log Parser GUI: Game Sessions + Vouchers/Notes + Cashouts + Rejected Handling
import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def cents_to_usd(cents):
    try:
        return f"${int(cents)/100:.2f}"
    except:
        return ""

def denom_to_usd(denom):
    try:
        return f"${int(denom)/100:.2f}"
    except:
        return ""

def to_12hr_format(time_str, tz_label="EST"):
    try:
        dt = datetime.strptime(time_str, "%H:%M:%S,%f")
        formatted = dt.strftime("%I:%M:%S,%f %p")[:-3]
        return f"{formatted} {tz_label}"
    except:
        return time_str

def format_duration(t1, t2):
    try:
        delta = t2 - t1
        total_seconds = int(delta.total_seconds())
        millis = int(delta.microseconds / 1000)
        return f"{total_seconds//3600:02}:{(total_seconds%3600)//60:02}:{total_seconds%60:02},{millis:03}"
    except:
        return ""

def auto_adjust_column_widths(writer, sheet_name):
    ws = writer.sheets[sheet_name]
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

def format_excel(path):
    try:
        wb = load_workbook(path)
        ws = wb.active

        # Auto column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        # Red + bold if Time Between Spins < 1s
        for row in range(2, ws.max_row + 1):
            val = ws[f"M{row}"].value
            if isinstance(val, str):
                match = re.match(r"(\d+):(\d+):(\d+),(\d+)", val)
                if match:
                    h, m, s, ms = map(int, match.groups())
                    total_ms = h * 3600000 + m * 60000 + s * 1000 + ms
                    if total_ms < 1000:
                        ws[f"M{row}"].font = Font(bold=True, color="FF0000")

        # Add borders
        from openpyxl.styles import Border, Side
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="medium", color="000000")

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                is_header = cell.row == 1
                is_edge = cell.col_idx == 1 or cell.col_idx == ws.max_column or is_header or cell.row == ws.max_row
                cell.border = Border(
                    left=thick if cell.col_idx == 1 else thin,
                    right=thick if cell.col_idx == ws.max_column else thin,
                    top=thick if is_header else thin,
                    bottom=thick if cell.row == ws.max_row else thin
                )

        wb.save(path)
    except Exception as e:
        print(f"Excel formatting skipped: {e}")


def parse_line(line):
    match = re.match(r'(\d{4}-\d{2}-\d{2}) (\d{2}:\d{2}:\d{2},\d{3}) (\w+)\s+([^|]+)\|\s+(.*)', line)
    if not match:
        return None
    date, time, msg_type, sender, message = match.groups()
    return {
        'Date': date,
        'Time': time,
        'MessageType': msg_type,
        'MessageSender': sender.strip(),
        'RawMessage': message.strip()
    }

def extract_key_values(message):
    data = {}
    if ':' in message:
        action, rest = message.split(':', 1)
        data['ActionType'] = action.strip().rstrip(':')
        parts = re.split(r'[; ]+', rest.strip())
    else:
        data['ActionType'] = message.strip()
        parts = []
    for part in parts:
        if '=' in part:
            key, val = part.split('=', 1)
            data[key.strip()] = val.strip()
    return data

def combine_all_fields(line):
    base = parse_line(line)
    if not base:
        return None
    extracted = extract_key_values(base['RawMessage'])
    base.update(extracted)
    return base

def browse_folder():
    folder_selected = filedialog.askdirectory()
    if not folder_selected:
        return

    tz_label = timezone_combo.get()
    messagebox.showinfo("Selected Timezone", f"Parsing logs with timezone: {tz_label}")

    for file in os.listdir(folder_selected):
        if file.endswith(".txt"):
            path = os.path.join(folder_selected, file)
            base = os.path.splitext(file)[0]
            with open(path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            raw_rows, summary_rows = [], []
            last_game_end = None
            i = 0

            while i < len(lines):
                line = lines[i]
                entry = combine_all_fields(line)
                if not entry:
                    i += 1
                    continue
                raw_rows.append(entry)
                action = entry.get("ActionType", "")

                if action == "--Beginning game":
                    game = {
                        "Date": entry["Date"],
                        "Title": entry.get("title", ""),
                        "Denom": denom_to_usd(entry.get("denom", "")),
                        "# of Lines": "",
                        "Bets Per Line": "",
                        "GameStart": to_12hr_format(entry["Time"], tz_label),
                        "Starting Balance": "",
                        "Bet Amount": "",
                        "Win Amount": "",
                        "Ending Balance": "",
                        "GameEnd": "",
                        "Length of Game": "",
                        "Time Between Spins": "",
                        "Action Type": ""
                    }
                    for j in range(i-1, -1, -1):
                        prev = combine_all_fields(lines[j])
                        if prev and prev.get("ActionType") == "Meters summary":
                            game["Starting Balance"] = cents_to_usd(prev.get("CurrentPlayableAmount", 0))
                            break
                    while i < len(lines):
                        sub = combine_all_fields(lines[i])
                        if not sub:
                            i += 1
                            continue
                        sub_action = sub.get("ActionType", "")
                        if sub_action == "End of game":
                            game["# of Lines"] = sub.get("#lines", "")
                            game["Bets Per Line"] = sub.get("bet_per_line", "")
                        elif sub_action == "sasEngine.gameEnd":
                            game["Bet Amount"] = cents_to_usd(sub.get("amountWagered", 0))
                            game["Win Amount"] = cents_to_usd(sub.get("amountWon", 0))
                            game["GameEnd"] = to_12hr_format(sub["Time"], tz_label)
                            for k in range(i+1, len(lines)):
                                post = combine_all_fields(lines[k])
                                if post and post.get("ActionType") == "Meters summary":
                                    game["Ending Balance"] = cents_to_usd(post.get("CurrentPlayableAmount", 0))
                                    break
                            try:
                                t1 = datetime.strptime(entry["Time"], "%H:%M:%S,%f")
                                t2 = datetime.strptime(sub["Time"], "%H:%M:%S,%f")
                                game["Length of Game"] = format_duration(t1, t2)
                                if last_game_end:
                                    game["Time Between Spins"] = format_duration(last_game_end, t1)
                                last_game_end = t2
                            except:
                                pass
                            summary_rows.append(game)
                            break
                        i += 1

                elif action == "Cashout initiated.":
                    cash = {
                        "Date": entry["Date"], "Title": "", "Denom": "",
                        "# of Lines": "", "Bets Per Line": "",
                        "GameStart": to_12hr_format(entry["Time"], tz_label),
                        "Starting Balance": "", "Bet Amount": "", "Win Amount": "",
                        "Ending Balance": "$0.00", "GameEnd": "",
                        "Length of Game": "", "Time Between Spins": "",
                        "Action Type": ""
                    }
                    for j in range(i-1, -1, -1):
                        prev = combine_all_fields(lines[j])
                        if prev and prev.get("ActionType") == "Meters summary":
                            cash["Starting Balance"] = cents_to_usd(prev.get("CurrentPlayableAmount", 0))
                            break
                    amt, val = "", ""
                    for j in range(i+1, len(lines)):
                        post = combine_all_fields(lines[j])
                        if post:
                            if post.get("ActionType") == "SAS TicketOut request":
                                amt = cents_to_usd(post.get("amt", 0))
                            elif post.get("ActionType") == "SAS TicketOut response - Success":
                                val = post.get("validation", "")
                            elif post.get("ActionType") == "Cashout complete.":
                                cash["GameEnd"] = to_12hr_format(post["Time"], tz_label)
                                cash["Action Type"] = f"{amt} Voucher Cashout, Val-ID: {val}"
                                break
                    summary_rows.append(cash)

                elif action in ["Ticket inserted", "Note inserted"]:
                    kind = "Voucher" if "Ticket" in action else "Bill"
                    val_id = entry.get("validation#", "")
                    start = to_12hr_format(entry["Time"], tz_label)
                    end, value = "", ""
                    status = ""
                    for j in range(i+1, len(lines)):
                        post = combine_all_fields(lines[j])
                        if post:
                            if kind == "Voucher" and "Ticket accepted" in post.get("ActionType", ""):
                                value = cents_to_usd(post.get("value", 0))
                                end = to_12hr_format(post["Time"], tz_label)
                                status = "Accepted"
                                break
                            elif kind == "Voucher" and "Ticket rejected" in post.get("ActionType", ""):
                                end = to_12hr_format(post["Time"], tz_label)
                                status = "Rejected"
                                break
                            elif kind == "Bill" and "Note accepted" in post.get("ActionType", ""):
                                value = cents_to_usd(post.get("valueCents", 0))
                                end = to_12hr_format(post["Time"], tz_label)
                                status = "Accepted"
                                break
                            elif kind == "Bill" and "Note rejected" in post.get("ActionType", ""):
                                end = to_12hr_format(post["Time"], tz_label)
                                status = "Rejected"
                                break
                    start_bal = ""
                    for j in range(i-1, -1, -1):
                        prev = combine_all_fields(lines[j])
                        if prev and prev.get("ActionType") == "Meters summary":
                            start_bal = cents_to_usd(prev.get("CurrentPlayableAmount", 0))
                            break
                    action_type = f"{value} {kind} Inserted/{status}"
                    if kind == "Voucher" and val_id:
                        action_type += f", Val-ID: {val_id}"
                    summary_rows.append({
                        "Date": entry["Date"], "Title": "", "Denom": "",
                        "# of Lines": "", "Bets Per Line": "",
                        "GameStart": start, "Starting Balance": start_bal,
                        "Bet Amount": "", "Win Amount": "",
                        "Ending Balance": value if status == "Accepted" else start_bal,
                        "GameEnd": end, "Length of Game": "", "Time Between Spins": "",
                        "Action Type": action_type
                    })
                i += 1

            df_raw = pd.DataFrame(raw_rows)
            df_summary = pd.DataFrame(summary_rows)
            base_out = os.path.join(folder_selected, base)
            df_raw.to_csv(base_out + "_Raw Extraction.csv", index=False)
            df_summary.to_csv(base_out + "_GameSummary.csv", index=False)
            with pd.ExcelWriter(base_out + "_GameSummary.xlsx", engine="openpyxl") as writer:
                df_summary.to_excel(writer, sheet_name="Game Summary", index=False)
                auto_adjust_column_widths(writer, "Game Summary")
            format_excel(base_out + "_GameSummary.xlsx")

    messagebox.showinfo("Done", f"Log parsing complete. Files saved to: {folder_selected}")

root = tk.Tk()
root.title("Log Parser Tool")
root.geometry("450x250")

tk.Label(root, text="Select a folder containing log .txt files:", pady=10).pack()
tk.Button(root, text="Browse Folder", command=browse_folder, height=2, width=20).pack(pady=10)
tk.Label(root, text="Select log timezone:").pack()
timezone_combo = ttk.Combobox(root, values=["EST", "CST", "MST", "PST", "UTC"], state="readonly")
timezone_combo.current(0)
timezone_combo.pack(pady=5)

root.mainloop()
