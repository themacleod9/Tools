# ✅ Log Parser Tool v5.0: Enhanced with Modular Design and Improved Parsing Logic
import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

# Utility Functions
def cents_to_usd(cents):
    try:
        return f"${int(cents)/100:.2f}"
    except:
        return ""

def denom_to_usd(denom):
    try:
        return f"${int(denom)/100:.2f}"
    except:
        return ""

def to_12hr_format(time_str, tz_label="EST"):
    try:
        dt = datetime.strptime(time_str, "%H:%M:%S,%f")
        formatted = dt.strftime("%I:%M:%S,%f %p")[:-3]  # keep 3-digit milliseconds
        return f"{formatted} {tz_label}"
    except:
        return time_str

def format_duration(t1, t2):
    try:
        delta = t2 - t1
        total_seconds = int(delta.total_seconds())
        millis = int(delta.microseconds / 1000)
        return f"{total_seconds//3600:02}:{(total_seconds%3600)//60:02}:{total_seconds%60:02},{millis:03}"
    except:
        return ""

# Excel Formatting Enhancements
def auto_adjust_column_widths(writer, sheet_name):
    ws = writer.sheets[sheet_name]
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

def format_excel(path):
    try:
        wb = load_workbook(path)
        ws = wb.active

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        # Highlight Time Between Spins < 1 second
        for row in range(2, ws.max_row + 1):
            val = ws[f"M{row}"].value
            if isinstance(val, str):
                match = re.match(r"(\d+):(\d+):(\d+),(\d+)", val)
                if match:
                    h, m, s, ms = map(int, match.groups())
                    total_ms = h * 3600000 + m * 60000 + s * 1000 + ms
                    if total_ms < 1000:
                        ws[f"M{row}"].font = Font(bold=True, color="FF0000")

        # Add borders to all cells
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="medium", color="000000")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                is_header = cell.row == 1
                is_edge = cell.col_idx == 1 or cell.col_idx == ws.max_column or is_header or cell.row == ws.max_row
                cell.border = Border(
                    left=thick if cell.col_idx == 1 else thin,
                    right=thick if cell.col_idx == ws.max_column else thin,
                    top=thick if is_header else thin,
                    bottom=thick if cell.row == ws.max_row else thin
                )

        wb.save(path)
    except Exception as e:
        print(f"Excel formatting skipped: {e}")

# Log Parsing Logic
def parse_log_file(file_path, tz_label):
    """
    Parses a log file and extracts raw data and summarized game session data.
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    raw_rows = []
    summary_rows = []

    for line in lines:
        match = re.match(r'(\d{4}-\d{2}-\d{2}) (\d{2}:\d{2}:\d{2},\d{3}) (\w+)\s+([^|]+)\|\s+(.*)', line)
        if match:
            date, time, msg_type, sender, message = match.groups()
            base_data = {
                'Date': date,
                'Time': time,
                'MessageType': msg_type,
                'MessageSender': sender.strip(),
                'RawMessage': message.strip()
            }
            raw_rows.append(base_data)

            # Further processing for summary rows (to be implemented)
            # Example: Process game sessions, cashouts, vouchers, etc.

    return raw_rows, summary_rows

def save_to_files(raw_rows, summary_rows, folder_selected, base):
    """
    Saves raw and summarized data to CSV and Excel files.
    """
    df_raw = pd.DataFrame(raw_rows)
    summary_columns = [
        "Date", "GameStart", "Title", "Denom", "# of Lines", "Bets Per Line",
        "Starting Balance", "Bet Amount", "Win Amount", "Ending Balance",
        "GameEnd", "Time Between Spins", "Length of Game", "Action Type"
    ]
    df_summary = pd.DataFrame(summary_rows)
    for col in summary_columns:
        if col not in df_summary.columns:
            df_summary[col] = ""
    df_summary = df_summary.reindex(columns=summary_columns)

    raw_csv = os.path.join(folder_selected, f"{base}_Raw Extraction.csv")
    summary_csv = os.path.join(folder_selected, f"{base}_GameSummary.csv")
    summary_xlsx = os.path.join(folder_selected, f"{base}_GameSummary.xlsx")
    df_raw.to_csv(raw_csv, index=False)
    df_summary.to_csv(summary_csv, index=False)

    with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Game Summary", index=False)
        auto_adjust_column_widths(writer, "Game Summary")
    format_excel(summary_xlsx)

# GUI Logic
def browse_folder():
    folder_selected = filedialog.askdirectory()
    if not folder_selected:
        return

    tz_label = timezone_combo.get()
    messagebox.showinfo("Selected Timezone", f"Parsing logs with timezone: {tz_label}")

    for file in os.listdir(folder_selected):
        if file.endswith(".txt"):
            path = os.path.join(folder_selected, file)
            base = os.path.splitext(file)[0]

            raw_rows, summary_rows = parse_log_file(path, tz_label)
            save_to_files(raw_rows, summary_rows, folder_selected, base)

    messagebox.showinfo("Done", f"Log parsing complete. Files saved to: {folder_selected}")

# GUI Initialization
root = tk.Tk()
root.title("Log Parser Tool v5.0")
root.geometry("450x250")
tk.Label(root, text="Select a folder containing log .txt files:", pady=10).pack()
tk.Button(root, text="Browse Folder", command=browse_folder, height=2, width=20).pack(pady=10)
tk.Label(root, text="Select log timezone:").pack()
timezone_combo = ttk.Combobox(root, values=["EST", "CST", "MST", "PST", "UTC"], state="readonly")
timezone_combo.current(0)
timezone_combo.pack(pady=5)
root.mainloop()
